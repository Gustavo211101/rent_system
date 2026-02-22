from __future__ import annotations

import calendar
from dataclasses import dataclass
from datetime import date, timedelta

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, Permission
from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden, Http404
from django.shortcuts import get_object_or_404, redirect, render

from .permissions import can_manage_staff

from events.models import Event

User = get_user_model()


@dataclass(frozen=True)
class _WeekSegment:
    event: Event
    start_col: int  # 0..6 (Mon..Sun)
    span: int  # 1..7
    lane: int  # 0..N


def _role_permission_choices():
    """Список прав, которые редактируем в UI ролей.

    Раньше в проекте в форме роли были чекбоксы. В шаблоне
    accounts/staff_role_form.html они всё ещё ожидаются
    (perm_choices/current_perm_ids). В какой-то момент обработчик
    прав в staff_role_*_view выпал — из-за этого чекбоксы пропали.

    Здесь мы показываем только "наши" кастомные права, чтобы
    интерфейс не превратился в простыню всех Permission Django.
    """

    wanted = [
        ("accounts", "manage_staff"),
        ("inventory", "manage_inventory"),
        ("events", "edit_event_card"),
        ("events", "edit_event_equipment"),
    ]

    qs = Permission.objects.select_related("content_type").filter(
        Q(content_type__app_label="accounts", codename="manage_staff")
        | Q(content_type__app_label="inventory", codename="manage_inventory")
        | Q(content_type__app_label="events", codename="edit_event_card")
        | Q(content_type__app_label="events", codename="edit_event_equipment")
    )

    # Сортируем в порядке wanted, чтобы список был стабильным
    order = {f"{a}.{c}": i for i, (a, c) in enumerate(wanted)}
    perms = sorted(
        list(qs),
        key=lambda p: order.get(f"{p.content_type.app_label}.{p.codename}", 999),
    )

    choices = []
    for p in perms:
        label = f"{p.content_type.app_label}: {p.name}"
        choices.append((int(p.id), label))
    return choices


def _deny():
    return HttpResponseForbidden("Недостаточно прав")


@login_required
def staff_users_view(request):
    if not can_manage_staff(request.user):
        return _deny()

    q = (request.GET.get("q") or "").strip()
    role = (request.GET.get("role") or "").strip()

    # Активные сверху, заблокированные ниже
    users_qs = User.objects.all().order_by("-is_active", "username")
    roles = Group.objects.all().order_by("name")

    if q:
        users_qs = users_qs.filter(
            Q(username__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
        )

    if role.isdigit():
        users_qs = users_qs.filter(groups__id=int(role)).distinct()

    return render(
        request,
        "accounts/staff_users.html",
        {"tab": "users", "users": users_qs, "q": q, "roles": roles, "role": role},
    )


def _month_grid(year: int, month: int) -> list[list[date]]:
    """Сетка месяца: недели (Пн..Вс), включая дни соседних месяцев."""
    cal = calendar.Calendar(firstweekday=calendar.MONDAY)
    weeks = []
    for week in cal.monthdatescalendar(year, month):
        weeks.append(list(week))
    return weeks


def _ru_month_name(month: int) -> str:
    names = [
        "Январь",
        "Февраль",
        "Март",
        "Апрель",
        "Май",
        "Июнь",
        "Июль",
        "Август",
        "Сентябрь",
        "Октябрь",
        "Ноябрь",
        "Декабрь",
    ]
    if 1 <= month <= 12:
        return names[month - 1]
    return ""


def _assign_lanes(segments: list[_WeekSegment]) -> list[_WeekSegment]:
    """Greedy lane assignment per week to avoid overlaps."""
    # Sort by start, then by longer spans first (slightly nicer)
    ordered = sorted(segments, key=lambda s: (s.start_col, -s.span, getattr(s.event, "id", 0)))
    lane_ends: list[int] = []  # last occupied col (inclusive) per lane
    out: list[_WeekSegment] = []
    for s in ordered:
        placed_lane = None
        for i, end_col in enumerate(lane_ends):
            if s.start_col > end_col:
                placed_lane = i
                lane_ends[i] = s.start_col + s.span - 1
                break
        if placed_lane is None:
            placed_lane = len(lane_ends)
            lane_ends.append(s.start_col + s.span - 1)
        out.append(_WeekSegment(event=s.event, start_col=s.start_col, span=s.span, lane=placed_lane))
    # Keep original order stable-ish for rendering
    return sorted(out, key=lambda s: (s.lane, s.start_col, -s.span))


@login_required
def staff_personnel_availability_calendar_view(request):
    """Календарь занятости сотрудников по мероприятиям."""
    if not can_manage_staff(request.user):
        return _deny()

    # Users list
    users = User.objects.filter(is_active=True).order_by("username")

    # Selected params
    user_id_raw = (request.GET.get("user_id") or "").strip()
    today = date.today()
    year = int((request.GET.get("year") or today.year) or today.year)
    month = int((request.GET.get("month") or today.month) or today.month)
    if month < 1:
        month = 1
    if month > 12:
        month = 12

    selected_user = None
    if user_id_raw.isdigit():
        selected_user = User.objects.filter(id=int(user_id_raw)).first()

    month_days = _month_grid(year, month)

    week_segments: dict[date, list[_WeekSegment]] = {week[0]: [] for week in month_days}
    week_lanes: dict[date, list[None]] = {week[0]: [None] * 1 for week in month_days}

    if selected_user:
        # Find events where the user participates.
        # Exclude deleted/cancelled. Keep draft/confirmed/closed for history.
        qs = (
            Event.objects.filter(is_deleted=False)
            .exclude(status=Event.STATUS_CANCELLED)
            .filter(
                Q(responsible=selected_user)
                | Q(s_engineer=selected_user)
                | Q(engineers=selected_user)
                | Q(role_slots__users=selected_user)
            )
            .distinct()
        )

        # Limit by month span (for performance)
        month_start = month_days[0][0]
        month_end = month_days[-1][-1]
        qs = qs.filter(start_date__lte=month_end, end_date__gte=month_start)
        events = list(qs.order_by("start_date", "id"))

        for week in month_days:
            ws = week[0]
            we = week[-1]
            segs: list[_WeekSegment] = []
            for ev in events:
                ev_start = ev.start_date
                ev_end = ev.end_date or ev.start_date
                if ev_end < ws or ev_start > we:
                    continue
                seg_start = max(ev_start, ws)
                seg_end = min(ev_end, we)
                start_col = (seg_start - ws).days
                span = (seg_end - seg_start).days + 1
                segs.append(_WeekSegment(event=ev, start_col=start_col, span=span, lane=0))

            assigned = _assign_lanes(segs)
            week_segments[ws] = assigned
            lanes_count = 1
            if assigned:
                lanes_count = max(s.lane for s in assigned) + 1
            # Template uses |length on this value (see template)
            week_lanes[ws] = [None] * lanes_count

    return render(
        request,
        "accounts/personnel_availability_calendar.html",
        {
            "tab": "users",
            "users": users,
            "selected_user": selected_user,
            "year": year,
            "month": month,
            "month_name": _ru_month_name(month),
            "month_days": month_days,
            "week_segments": week_segments,
            "week_lanes": week_lanes,
        },
    )


@login_required
def staff_roles_view(request):
    if not can_manage_staff(request.user):
        return _deny()

    roles = Group.objects.all().order_by("name")
    return render(request, "accounts/staff_roles.html", {"tab": "roles", "roles": roles})


@login_required
def staff_user_add_view(request):
    if not can_manage_staff(request.user):
        return _deny()

    roles = Group.objects.all().order_by("name")

    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = (request.POST.get("password") or "").strip()
        first_name = (request.POST.get("first_name") or "").strip()
        last_name = (request.POST.get("last_name") or "").strip()
        email = (request.POST.get("email") or "").strip()

        # ✅ много ролей
        role_ids = request.POST.getlist("roles")
        # совместимость: если вдруг где-то осталось старое поле role_id
        role_id_single = (request.POST.get("role_id") or "").strip()

        if not username or not password:
            messages.error(request, "Нужны username и пароль.")
            return render(
                request,
                "accounts/staff_user_form.html",
                {
                    "tab": "users",
                    "roles": roles,
                    "selected_role_ids": role_ids,
                },
            )

        if User.objects.filter(username=username).exists():
            messages.error(request, "Пользователь с таким username уже существует.")
            return render(
                request,
                "accounts/staff_user_form.html",
                {
                    "tab": "users",
                    "roles": roles,
                    "selected_role_ids": role_ids,
                },
            )

        user = User.objects.create_user(username=username, password=password)
        if hasattr(user, "first_name"):
            user.first_name = first_name
        if hasattr(user, "last_name"):
            user.last_name = last_name
        if hasattr(user, "email"):
            user.email = email
        user.save()

        # ✅ сохраняем группы
        user.groups.clear()

        chosen = []
        for rid in role_ids:
            if str(rid).isdigit():
                chosen.append(int(rid))

        if not chosen and role_id_single.isdigit():
            chosen = [int(role_id_single)]

        if chosen:
            user.groups.add(*Group.objects.filter(id__in=chosen))

        messages.success(request, "Пользователь создан.")
        return redirect("staff_users")

    return render(
        request,
        "accounts/staff_user_form.html",
        {
            "tab": "users",
            "roles": roles,
            "selected_role_ids": [],
        },
    )


@login_required
def staff_user_edit_view(request, user_id: int):
    if not can_manage_staff(request.user):
        return _deny()

    user_obj = get_object_or_404(User, id=user_id)
    roles = Group.objects.all().order_by("name")

    if getattr(user_obj, "is_superuser", False):
        return HttpResponseForbidden("Суперпользователя редактируем только через админку.")

    if request.method == "POST":
        first_name = (request.POST.get("first_name") or "").strip()
        last_name = (request.POST.get("last_name") or "").strip()
        email = (request.POST.get("email") or "").strip()

        # ✅ много ролей
        role_ids = request.POST.getlist("roles")
        role_id_single = (request.POST.get("role_id") or "").strip()

        new_password = (request.POST.get("password") or "").strip()

        if hasattr(user_obj, "first_name"):
            user_obj.first_name = first_name
        if hasattr(user_obj, "last_name"):
            user_obj.last_name = last_name
        if hasattr(user_obj, "email"):
            user_obj.email = email

        if new_password:
            user_obj.set_password(new_password)

        user_obj.groups.clear()

        chosen = []
        for rid in role_ids:
            if str(rid).isdigit():
                chosen.append(int(rid))

        if not chosen and role_id_single.isdigit():
            chosen = [int(role_id_single)]

        if chosen:
            user_obj.groups.add(*Group.objects.filter(id__in=chosen))

        user_obj.save()
        messages.success(request, "Пользователь обновлён.")
        return redirect("staff_users")

    selected_role_ids = [str(g.id) for g in user_obj.groups.all()]

    return render(
        request,
        "accounts/staff_user_form.html",
        {
            "tab": "users",
            "user_obj": user_obj,
            "roles": roles,
            "selected_role_ids": selected_role_ids,
        },
    )


@login_required
def staff_user_delete_view(request, user_id: int):
    """
    ВАЖНО: пользователей НЕ удаляем физически.
    Вместо удаления — блокировка is_active=False (повторно — разблокировка).
    """
    if not can_manage_staff(request.user):
        return _deny()

    user_obj = get_object_or_404(User, id=user_id)

    if getattr(user_obj, "is_superuser", False):
        return HttpResponseForbidden("Суперпользователя блокируем только через админку.")

    if user_obj.id == request.user.id:
        return HttpResponseForbidden("Нельзя заблокировать самого себя.")

    is_active = getattr(user_obj, "is_active", True)
    action = "block" if is_active else "unblock"

    if request.method == "POST":
        user_obj.is_active = (action == "unblock")
        user_obj.save(update_fields=["is_active"])

        if action == "block":
            messages.success(request, "Пользователь заблокирован.")
        else:
            messages.success(request, "Пользователь разблокирован.")
        return redirect("staff_users")

    if action == "block":
        title = "Блокировка"
        confirm_text = "Точно заблокировать"
        action_button_text = "Заблокировать"
    else:
        title = "Разблокировка"
        confirm_text = "Точно разблокировать"
        action_button_text = "Разблокировать"

    return render(
        request,
        "accounts/staff_confirm_delete.html",
        {
            "tab": "users",
            "title": title,
            "confirm_text": confirm_text,
            "action_button_text": action_button_text,
            "user_obj": user_obj,
            "cancel_url": "staff_users",
        },
    )


@login_required
def staff_role_add_view(request):
    if not can_manage_staff(request.user):
        return _deny()

    perm_choices = _role_permission_choices()

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        if not name:
            messages.error(request, "Название роли обязательно.")
            return render(
                request,
                "accounts/staff_role_form.html",
                {"tab": "roles", "perm_choices": perm_choices, "current_perm_ids": []},
            )

        if Group.objects.filter(name=name).exists():
            messages.error(request, "Такая роль уже существует.")
            return render(
                request,
                "accounts/staff_role_form.html",
                {"tab": "roles", "perm_choices": perm_choices, "current_perm_ids": []},
            )

        role = Group.objects.create(name=name)

        # ✅ сохраняем права роли из чекбоксов
        perm_ids = [p for p in request.POST.getlist("perm_ids") if str(p).isdigit()]
        if perm_ids:
            role.permissions.set(Permission.objects.filter(id__in=[int(x) for x in perm_ids]))

        messages.success(request, "Роль создана.")
        return redirect("staff_roles")

    return render(
        request,
        "accounts/staff_role_form.html",
        {"tab": "roles", "perm_choices": perm_choices, "current_perm_ids": []},
    )


@login_required
def staff_role_edit_view(request, group_id: int):
    if not can_manage_staff(request.user):
        return _deny()

    role = get_object_or_404(Group, id=group_id)

    perm_choices = _role_permission_choices()
    current_perm_ids = [int(p.id) for p in role.permissions.all()]

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        if not name:
            messages.error(request, "Название роли обязательно.")
            return render(
                request,
                "accounts/staff_role_form.html",
                {
                    "tab": "roles",
                    "role": role,
                    "perm_choices": perm_choices,
                    "current_perm_ids": current_perm_ids,
                },
            )

        if Group.objects.filter(name=name).exclude(id=role.id).exists():
            messages.error(request, "Такая роль уже существует.")
            return render(
                request,
                "accounts/staff_role_form.html",
                {
                    "tab": "roles",
                    "role": role,
                    "perm_choices": perm_choices,
                    "current_perm_ids": current_perm_ids,
                },
            )

        role.name = name
        role.save()

        # ✅ сохраняем права роли из чекбоксов
        perm_ids = [p for p in request.POST.getlist("perm_ids") if str(p).isdigit()]
        role.permissions.set(Permission.objects.filter(id__in=[int(x) for x in perm_ids]))

        messages.success(request, "Роль обновлена.")
        return redirect("staff_roles")

    return render(
        request,
        "accounts/staff_role_form.html",
        {
            "tab": "roles",
            "role": role,
            "perm_choices": perm_choices,
            "current_perm_ids": current_perm_ids,
        },
    )


@login_required
def staff_role_delete_view(request, group_id: int):
    if not can_manage_staff(request.user):
        return _deny()

    role = get_object_or_404(Group, id=group_id)

    if User.objects.filter(groups=role).exists():
        if request.method == "POST":
            messages.error(request, "Нельзя удалить роль: она назначена пользователям.")
            return redirect("staff_roles")
        return render(
            request,
            "accounts/staff_role_confirm_delete.html",
            {"tab": "roles", "role": role, "has_users": True},
        )

    if request.method == "POST":
        role.delete()
        messages.success(request, "Роль удалена.")
        return redirect("staff_roles")

    return render(
        request,
        "accounts/staff_role_confirm_delete.html",
        {"tab": "roles", "role": role, "has_users": False},
    )


# ============================================================
# ✅ ДОБАВЛЕНО: просмотр анкеты сотрудника + экспорт анкеты в Excel
# ============================================================

def _get_related_profile_object(user_obj):
    """
    Пытаемся найти связанную "анкету/профиль" пользователя в проекте.
    Ничего не ломаем: если модели нет или связь отсутствует — вернём None.
    """
    candidates = [
        "profile",
        "questionnaire",
        "form",
        "staff_profile",
        "employee_profile",
        "personnel_profile",
        " анкета",  # на случай экзотики
    ]
    for attr in candidates:
        try:
            obj = getattr(user_obj, attr, None)
        except Exception:
            obj = None
        if obj is not None:
            return obj
    return None


def _model_to_pairs(model_obj):
    """
    Любую модель Django превращаем в список (label, value) по ВСЕМ полям.
    FileField/ImageField отдаём как url/path (что есть).
    """
    pairs = []
    if model_obj is None:
        return pairs

    try:
        fields = model_obj._meta.fields
    except Exception:
        return pairs

    for f in fields:
        name = getattr(f, "name", "")
        if not name or name in ("id",):
            continue

        # Обычно FK на user называется user/owner/employee — пропускаем, чтобы не дублировать
        if name in ("user", "owner", "employee", "account"):
            continue

        label = getattr(f, "verbose_name", name)
        try:
            val = getattr(model_obj, name)
        except Exception:
            val = ""

        # FileField/ImageField
        try:
            if hasattr(val, "url"):
                val = val.url
            elif hasattr(val, "path"):
                val = val.path
        except Exception:
            pass

        # Datetime/date formatting
        try:
            if hasattr(val, "strftime"):
                # если это date/datetime
                val = val.strftime("%Y-%m-%d %H:%M") if "time" in str(type(val)).lower() else val.strftime("%Y-%m-%d")
        except Exception:
            pass

        pairs.append((str(label), "" if val is None else str(val)))
    return pairs


@login_required
def staff_user_view(request, user_id: int):
    """
    Просмотр анкеты/профиля сотрудника менеджером.
    Ожидаемый шаблон: accounts/staff_user_view.html
    """
    if not can_manage_staff(request.user):
        return _deny()

    user_obj = get_object_or_404(User, id=user_id)

    profile_obj = _get_related_profile_object(user_obj)
    user_pairs = _model_to_pairs(user_obj)
    profile_pairs = _model_to_pairs(profile_obj)

    return render(
        request,
        "accounts/staff_user_view.html",
        {
            "tab": "users",
            "user_obj": user_obj,
            "profile_obj": profile_obj,
            "user_pairs": user_pairs,
            "profile_pairs": profile_pairs,
        },
    )


@login_required
def staff_user_export_xlsx(request, user_id: int):
    """
    Экспорт ВСЕЙ анкеты сотрудника (User + связанная анкета) в Excel.
    """
    if not can_manage_staff(request.user):
        return _deny()

    user_obj = get_object_or_404(User, id=user_id)
    profile_obj = _get_related_profile_object(user_obj)

    # Собираем пары
    rows = []
    rows.append(("=== АККАУНТ ===", ""))
    rows.extend(_model_to_pairs(user_obj))

    # группы/роли отдельной строкой
    try:
        roles = ", ".join([g.name for g in user_obj.groups.all().order_by("name")])
        rows.append(("Роли (groups)", roles))
    except Exception:
        pass

    rows.append(("", ""))
    rows.append(("=== АНКЕТА ===", ""))

    if profile_obj is None:
        rows.append(("Анкета", "Не найдена (нет связанной модели профиля/анкеты)"))
    else:
        rows.extend(_model_to_pairs(profile_obj))

    # Генерация Excel
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise Http404(f"openpyxl не доступен: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Анкета"

    ws.append(["Поле", "Значение"])
    for a, b in rows:
        ws.append([a, b])

    # Автоширина
    for col in range(1, 3):
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 70)

    # отдаём файлом
    filename = f"profile_user_{user_obj.id}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp