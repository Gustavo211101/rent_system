from __future__ import annotations

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render

from accounts.permissions import can_edit_inventory, can_view_stock

from .warehouse_forms import StockCategoryForm, StockSubcategoryForm
from .models import StockCategory, StockSubcategory, StockEquipmentType, StockEquipmentItem

import openpyxl
from openpyxl import load_workbook
from .warehouse_forms import StockImportForm
from .warehouse_import import import_stock_from_rows


def _forbidden():
    return HttpResponseForbidden("Недостаточно прав")


@login_required
def stock_category_list_view(request):
    if not can_view_stock(request.user):
        return _forbidden()

    categories = StockCategory.objects.all().order_by("name")
    return render(
        request,
        "inventory/warehouse/categories/list.html",
        {
            "categories": categories,
            "can_manage": can_edit_inventory(request.user),
        },
    )


@login_required
def stock_category_add_view(request):
    if not can_edit_inventory(request.user):
        return _forbidden()

    if request.method == "POST":
        form = StockCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Категория добавлена")
            return redirect("stock_category_list")
        messages.error(request, "Исправьте ошибки в форме")
    else:
        form = StockCategoryForm()

    return render(
        request,
        "inventory/warehouse/categories/form.html",
        {
            "form": form,
            "title": "Добавить категорию",
        },
    )


@login_required
def stock_category_edit_view(request, category_id: int):
    if not can_edit_inventory(request.user):
        return _forbidden()

    category = get_object_or_404(StockCategory, id=category_id)

    if request.method == "POST":
        form = StockCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Категория обновлена")
            return redirect("stock_category_list")
        messages.error(request, "Исправьте ошибки в форме")
    else:
        form = StockCategoryForm(instance=category)

    return render(
        request,
        "inventory/warehouse/categories/form.html",
        {
            "form": form,
            "title": "Редактировать категорию",
            "category": category,
        },
    )


@login_required
def stock_category_delete_view(request, category_id: int):
    if not can_edit_inventory(request.user):
        return _forbidden()

    category = get_object_or_404(StockCategory, id=category_id)

    has_subcats = StockSubcategory.objects.filter(category=category).exists()
    has_types = StockEquipmentType.objects.filter(category=category).exists()
    has_items = StockEquipmentItem.objects.filter(equipment_type__category=category).exists()

    if request.method == "POST":
        if has_subcats or has_types or has_items:
            messages.error(request, "Нельзя удалить категорию: в ней есть подкатегории/оборудование")
            return redirect("stock_category_list")
        category.delete()
        messages.success(request, "Категория удалена")
        return redirect("stock_category_list")

    return render(
        request,
        "inventory/warehouse/categories/confirm_delete.html",
        {
            "category": category,
            "has_subcats": has_subcats,
            "has_items": (has_types or has_items),
        },
    )


@login_required
def stock_category_detail_view(request, category_id: int):
    if not can_view_stock(request.user):
        return _forbidden()

    category = get_object_or_404(StockCategory, id=category_id)
    subcategories = StockSubcategory.objects.filter(category=category).order_by("name")

    return render(
        request,
        "inventory/warehouse/subcategories/list.html",
        {
            "category": category,
            "subcategories": subcategories,
            "can_manage": can_edit_inventory(request.user),
        },
    )


@login_required
def stock_subcategory_list_view(request, category_id: int):
    """Alias: /warehouse/categories/<id>/subcategories/"""
    return stock_category_detail_view(request, category_id)


@login_required
def stock_subcategory_add_view(request, category_id: int):
    if not can_edit_inventory(request.user):
        return _forbidden()

    category = get_object_or_404(StockCategory, id=category_id)

    if request.method == "POST":
        form = StockSubcategoryForm(request.POST)
        if form.is_valid():
            subcat = form.save(commit=False)
            subcat.category = category
            subcat.save()
            messages.success(request, "Подкатегория добавлена")
            return redirect("stock_category_detail", category_id=category.id)
        messages.error(request, "Исправьте ошибки в форме")
    else:
        form = StockSubcategoryForm()

    return render(
        request,
        "inventory/warehouse/subcategories/form.html",
        {
            "form": form,
            "title": "Добавить подкатегорию",
            "category": category,
        },
    )


@login_required
def stock_subcategory_edit_view(request, category_id: int, subcategory_id: int):
    if not can_edit_inventory(request.user):
        return _forbidden()

    category = get_object_or_404(StockCategory, id=category_id)
    subcategory = get_object_or_404(StockSubcategory, id=subcategory_id, category=category)

    if request.method == "POST":
        form = StockSubcategoryForm(request.POST, instance=subcategory)
        if form.is_valid():
            form.save()
            messages.success(request, "Подкатегория обновлена")
            return redirect("stock_category_detail", category_id=category.id)
        messages.error(request, "Исправьте ошибки в форме")
    else:
        form = StockSubcategoryForm(instance=subcategory)

    return render(
        request,
        "inventory/warehouse/subcategories/form.html",
        {
            "form": form,
            "title": "Редактировать подкатегорию",
            "category": category,
            "subcategory": subcategory,
        },
    )


@login_required
def stock_subcategory_delete_view(request, category_id: int, subcategory_id: int):
    if not can_edit_inventory(request.user):
        return _forbidden()

    category = get_object_or_404(StockCategory, id=category_id)
    subcategory = get_object_or_404(StockSubcategory, id=subcategory_id, category=category)

    has_types = StockEquipmentType.objects.filter(subcategory=subcategory).exists()
    has_items = StockEquipmentItem.objects.filter(equipment_type__subcategory=subcategory).exists()

    if request.method == "POST":
        if has_types or has_items:
            messages.error(request, "Нельзя удалить подкатегорию: в ней есть оборудование")
            return redirect("stock_category_detail", category_id=category.id)
        subcategory.delete()
        messages.success(request, "Подкатегория удалена")
        return redirect("stock_category_detail", category_id=category.id)

    return render(
        request,
        "inventory/warehouse/subcategories/confirm_delete.html",
        {
            "subcategory": subcategory,
            "has_items": (has_types or has_items),
        },
    )

@login_required
def stock_import_view(request):
    if request.method == "POST":
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Файл не выбран.")
            return redirect("stock_import")

        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

        result = import_stock_from_rows(rows)
        return render(request, "inventory/warehouse/import.html", {"result": result})

    return render(request, "inventory/warehouse/import.html")